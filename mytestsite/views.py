from django.http import HttpResponse
from django.template import Template, Context
from django.shortcuts import render
from pathlib import Path
from copy import copy
from typing import Union, Optional
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def copy_excel_cell_range(
        src_ws: openpyxl.worksheet.worksheet.Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
        tgt_min_row: int = 1,
        tgt_min_col: int = 1,
        with_style: bool = True
) -> openpyxl.worksheet.worksheet.Worksheet:
    
    if tgt_ws is None:
        tgt_ws = src_ws

    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
        for cell in row:
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value
            )
            if with_style and cell.has_style:
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws


def append_df_to_excel(
        filename: Union[str, Path],
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1',
        startrow: Optional[int] = None,
        max_col_width: int = 30,
        autofilter: bool = False,
        fmt_int: str = "#,##0",
        fmt_float: str = "#,##0.00",
        fmt_date: str = "yyyy-mm-dd",
        fmt_datetime: str = "yyyy-mm-dd hh:mm",
        truncate_sheet: bool = False,
        storage_options: Optional[dict] = None,
        **to_excel_kwargs
) -> None:
    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt
    filename = Path(filename)
    file_exists = filename.is_file()
    first_col = int(to_excel_kwargs.get("index", True)) + 1
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
        filename.with_suffix(".xlsx"),
        engine="openpyxl",
        mode="a" if file_exists else "w",
        if_sheet_exists="new" if file_exists else None,
        date_format=fmt_date,
        datetime_format=fmt_datetime,
        storage_options=storage_options
    ) as writer:
        if file_exists:
            writer.book = wb
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                idx = writer.book.sheetnames.index(sheet_name)
                writer.book.remove(writer.book.worksheets[idx])
                writer.book.create_sheet(sheet_name, idx)
            writer.sheets = sheets
        else:
            startrow = 0

        df.T.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs, index= False)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
            col_no = xl_col_no - first_col
            width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                        df.columns[col_no]) + 6
            width = min(max_col_width, width)
            column_letter = get_column_letter(xl_col_no)
            worksheet.column_dimensions[column_letter].width = width
            if np.issubdtype(dtyp, np.integer):
                set_column_format(worksheet, column_letter, fmt_int)
            if np.issubdtype(dtyp, np.floating):
                set_column_format(worksheet, column_letter, fmt_float)

    if file_exists and sheet_exists:
        wb = load_workbook(filename)
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=startrow + 1,
            with_style=True
        )
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()

def main(request):
    
    plantilla_main=open("/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/templates/main.html")
    plt=Template(plantilla_main.read())
    plantilla_main.close()
    ctxto=Context()
    ventana_principal=plt.render(ctxto)
    return HttpResponse(ventana_principal)

def Ventana_KPIs(request):
    
    mensaje="Registro realizado:"
    datos_faltantes = request.POST.dict()
    if len(datos_faltantes) > 0:
       nLista_faltante = list(datos_faltantes.values())
       df = pd.DataFrame(nLista_faltante) 
       print("Datos guardados", df) 
       df.to_csv('indicadores.to_csv')
       dataset = pd.read_excel("/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_vias_mensual.xlsx" )
       dataset2 = pd.read_excel("/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_mtto_anual.xlsx" )
       dataset3 = pd.read_excel("/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_trenes_mensual.xlsx" )

       data = pd.read_csv('/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/indicadores.to_csv')
       data_faltante = pd.DataFrame(data)

       """ 
       VIAS 
       """

       #Pre-operaciones VIAS
       TPI_i = dataset[['TPI_i']].sum()    
       KA_i = dataset[['KA_i']].sum()  
       MTBF_Pond = float(data_faltante.iloc[2,1])/float(KA_i)
       Act_Mtto= dataset2.groupby(['Real_en'])['Real_en'].count() 
       Act_Mtto.to_frame()
       Mtto_VIAS = float(Act_Mtto.loc['Vias'])

       #Indicadores_VIAS
       D_v = float(1- (float(TPI_i))/float(data_faltante.iloc[1,1])) * 100
       F_v = float(MTBF_Pond)/float(data_faltante.iloc[3,1])        
       M_v = float(Mtto_VIAS)/float(data_faltante.iloc[4,1]) * 100


       """
       TRENES 
       """

       #Pre-operaciones TRENES
       Ave_Trenes= dataset3.groupby(['Tipo_Tren'])['Tipo_Tren'].count() 
       Ave_Trenes.to_frame()
       Ave_TN = Ave_Trenes.loc['Nuevo']
       Ave_TNM16 = Ave_Trenes.loc['NM16']
      # Mtto_Trenes = float(Act_Mtto.loc['Trenes'])
       AMR = dataset2.groupby(['Peso_Act'])['Peso_Act'].count() 
       AMR.to_frame()
       AMR_T1 = float(AMR.loc[1])
       AMR_T2 = float(AMR.loc[1.3])
       AMR_T3 = float(AMR.loc[1.7])
       AMR_T4 = float(AMR.loc[2])
       AMR_T5 = float(AMR.loc[2.3])
       
       Total_AMR = ((float(AMR_T1) * 1) + (float(AMR_T2) * 1.3) + (float(AMR_T3) * 1.7) + (float(AMR_T4) * 2) + (float(AMR_T5) * 2.3))
       AMP = ((float(data_faltante.iloc[7,1]) * 1) + (float(data_faltante.iloc[8,1]) * 1.3) + (float(data_faltante.iloc[9,1]) * 1.7) + (float(data_faltante.iloc[10,1]) * 2) + (float(data_faltante.iloc[11,1]) * 2.3))

     


       #Indicadores_Trenes
       F_TN = float(data_faltante.iloc[5,1])/float(Ave_TN)
       F_TNM16 = float(data_faltante.iloc[6,1])/float(Ave_TNM16)
       M_t = ((float(Total_AMR/AMP)* 100) - (float(data_faltante.iloc[12,1])))

       Valores_Indicadores_Calculados = {"Disponibilidad de via:": D_v, 
                                         "Fiabibilidad de via:": F_v,
                                         "Mantenimiento de via:": M_v, 
                                         "Fiabibilidad de trenes nuevos:": F_TN ,
                                         "Fiabibilidad de trenes NM16:": F_TNM16, 
                                         "Mantenimiento de trenes:": M_t }
                                      
       Indicadores_Calculados_df = pd.DataFrame(Valores_Indicadores_Calculados.items(),columns=["KPI","Valor"])
       Indicadores_Calculados_df.to_csv("Valores de KPIs.csv") 
               #aqui introducir código de graficas                                                 
       return render(request,'Ventana_KPIs.html')
    else: 
    
       return render(request,'Ventana_KPIs.html')



def formulario_trenes(request):   

    mensaje="Registro realizado:"
    datos = request.POST.dict()
    nLista = list(datos.values())
    df = pd.DataFrame(nLista) 
    print(mensaje, nLista) 
    filename = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_trenes_mensual.xlsx'
    filename2 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_trenes_anual.xlsx'
    filename3 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_trenes_historico.xlsx'
    append_df_to_excel(filename, df, header=False)
    append_df_to_excel(filename2, df, header=False)
    append_df_to_excel(filename3, df, header=False)
    return render(request,'formulario_trenes.html')


def formulario_SC(request):    
  
    mensaje="Registro realizado:"
    datos = request.POST.dict()
    nLista = list(datos.values())
    df = pd.DataFrame(nLista) 
    print(mensaje, nLista)  
    filename = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_SC_mensual.xlsx'
    filename2 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_SC_anual.xlsx'
    filename3 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_SC_historico.xlsx'
    append_df_to_excel(filename, df, header=False)
    append_df_to_excel(filename2, df, header=False)
    append_df_to_excel(filename3, df, header=False)
    return render(request,'formulario_SC.html')
    
    


def formulario_mtto(request):    
    
    mensaje="Registro realizado:"
    datos = request.POST.dict()
    nLista = list(datos.values())
    df = pd.DataFrame(nLista) 
    print(mensaje, nLista) 
    filename = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_mtto_mensual.xlsx'
    filename2 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_mtto_anual.xlsx'
    filename3 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_mtto_historico.xlsx'
    append_df_to_excel(filename, df, header=False)
    append_df_to_excel(filename2, df, header=False)
    append_df_to_excel(filename3, df, header=False)
    return render(request,'formulario_mtto.html')


def formulario_VIAS(request):    
    
    mensaje="Registro realizado:"
    datos = request.POST.dict()
    if len(datos) > 0:
       nLista = list(datos.values()) 
       print(nLista, "Hola Chayo")
       nColumne = str(nLista[4]) + str(nLista[5])
       nLista.insert(-1,nColumne)
       
       Ubic_AB = { 1:7.5,
                   2:2.5,
                   3:2.5,
                   4:2.5,
                5:2.5,
                6:2.5,
                7:7.5,
                8:5,
                9:2.5,
                10:7.5,
                11:5,
                12:7.5,
                13:7.5,
                14:5,
                15:5,
                16:2.5,
                17:5,
                18:5,
                19:7.5,
                20:7.5 }
       resultado  = 0
       pesos_ruta = []    
       for key in Ubic_AB:
          if int(nLista[4]) <= key <= int(nLista[5]):
             print(Ubic_AB[key], "el de enmedio")
             pesos_ruta.append(Ubic_AB[key])
       suma_pesos = sum(pesos_ruta)
       nLista.insert(9, suma_pesos)
       nPond = float(nLista[3])/int(600) * float(nLista[9]) #TPI_i #Ya esta en horas por acción de /60
       nLista.insert(8,nPond)
       df = pd.DataFrame(nLista) 
       #print(mensaje, nLista, nColumne, suma_pesos) 
       filename = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_vias_mensual.xlsx'
       filename2 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_vias_anual.xlsx'
       filename3 = r'/home/uranga/Documentos/STC-Py/AplicaciónDjango/mytestsite/mytestsite/registros_vias_historico.xlsx'
       append_df_to_excel(filename, df, header=False)
       append_df_to_excel(filename2, df, header=False)
       append_df_to_excel(filename3, df, header=False)
       return render(request,'formulario_VIAS.html')
      
    else: 
    
       return render(request,'formulario_VIAS.html')



    

